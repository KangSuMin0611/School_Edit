import urllib.request
from bs4 import BeautifulSoup as bs
from openpyxl import Workbook

def main():
    articles = []
    
    for i in range(1,32):
        articles.append("5월 "+ str(i)+"일 뉴스내용")
        if 1<=i and i<=9:
            url = 'http://news.sbs.co.kr/news/newsSection.do?sectionType=01&pageDate=2021050'+str(i)
        else:
            url = 'http://news.sbs.co.kr/news/newsSection.do?sectionType=01&pageDate=202105'+str(i)
            
        sourcecode = urllib.request.urlopen(url).read()
        soup = bs(sourcecode,'html.parser')
        div = soup.find("div", class_="w_news_list type_issue")
        for i in div.find_all("strong"):
            articles.append(i.get_text())
            
    #print(articles)
    wb = Workbook()
    sheet1 = wb.active
    file_name = 'sbs.xlsx'
    
    for i in range(0, len(articles)):
        sheet1.cell(row=i+1,column = 1).value = i+1
        sheet1.cell(row=i+1,column = 2).value = articles[i]
    wb.save(filename=file_name)
    
    
if __name__ == "__main__":
    main()